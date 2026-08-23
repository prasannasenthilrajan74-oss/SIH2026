"""
Export MPLADS Sentinel showcase dataset to a formatted Excel workbook.
Output: Dataset/MPLADS_Sentinel_Showcase_Dataset.xlsx
"""
import sys, os
sys.path.insert(0, ".")
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from backend.app.db.session import SessionLocal
from backend.app.models.models import Work, Document, Agency, RiskScore, Payment

db = SessionLocal()
print("Fetching data...")
works = db.query(Work).all()
agencies = db.query(Agency).all()
docs = db.query(Document).all()
payments = db.query(Payment).all()

proj_rows = []
for w in works:
    rs = w.risk_scores
    score = round(rs.overall_score, 1) if rs else 0.0
    proj_rows.append({
        "Project ID": w.id, "Description": w.description, "Category": w.category,
        "MP Name": w.mp_name, "Constituency": w.constituency, "State": w.state_code,
        "District Code": w.district_code, "Status": w.status,
        "Sanctioned Amount (Rs)": w.sanctioned_amount, "Expenditure (Rs)": w.expenditure,
        "Physical Progress (%)": w.physical_progress, "Financial Progress (%)": w.financial_progress,
        "Sanction Date": str(w.sanction_date) if w.sanction_date else "",
        "Expected Completion": str(w.expected_completion_date) if w.expected_completion_date else "",
        "Actual Completion": str(w.actual_completion_date) if w.actual_completion_date else "",
        "Latitude": w.latitude, "Longitude": w.longitude,
        "Risk Score": score,
        "Financial Risk": round(rs.financial_risk, 1) if rs else 0.0,
        "Delay Risk": round(rs.delay_risk, 1) if rs else 0.0,
        "Cost Risk": round(rs.cost_risk, 1) if rs else 0.0,
        "Duplicate Risk": round(rs.duplicate_risk, 1) if rs else 0.0,
        "Document Risk": round(rs.document_risk, 1) if rs else 0.0,
        "Risk Level": ("Critical" if score >= 80 else "High" if score >= 60 else "Medium" if score >= 40 else "Low"),
    })

ag_rows = []
for ag in agencies:
    cr = ag.completion_rate or 0; cd = ag.average_cost_deviation or 0; rs = round(ag.risk_score or 0, 1)
    ag_rows.append({
        "Agency ID": ag.id, "Agency Name": ag.name, "District": ag.district_code,
        "Risk Score": rs, "Risk Level": ("Critical" if rs >= 80 else "High" if rs >= 60 else "Medium" if rs >= 40 else "Low"),
        "Completion Rate (%)": round((cr*100) if cr <= 1.0 else cr, 1),
        "Avg Delay Days": round(ag.average_delay_days or 0, 1),
        "Avg Cost Deviation (%)": round((cd*100) if abs(cd) <= 1.0 else cd, 1),
    })

doc_rows = []
for d in docs:
    ed = d.extracted_data or {}
    doc_rows.append({
        "Document ID": d.id, "Work ID": d.work_id, "Document Type": d.document_type,
        "File Name": d.file_name, "Consistency Score": d.consistency_score,
        "Vendor": ed.get("vendor",""), "Sanctioned Amount": ed.get("sanctioned_amount",""),
        "Extracted Amount": ed.get("extracted_amount",""),
        "Mismatch Flag": "YES" if ed.get("mismatch") else "NO", "Category": ed.get("category",""),
    })

pay_rows = []
for p in payments:
    pay_rows.append({
        "Transaction ID": p.id, "Work ID": p.work_id, "Payment Type": p.payment_type,
        "Amount (Rs)": p.amount, "Payment Date": str(p.payment_date) if p.payment_date else "",
        "Transaction Ref": p.transaction_ref,
    })

scores   = [r["Risk Score"] for r in proj_rows]
mm_count = sum(1 for r in doc_rows if r["Mismatch Flag"] == "YES")
summary_rows = [
    ("Total Projects", len(works)),
    ("  Critical (>=80)", len([s for s in scores if s >= 80])),
    ("  High (60-80)",    len([s for s in scores if 60 <= s < 80])),
    ("  Medium (40-60)",  len([s for s in scores if 40 <= s < 60])),
    ("  Low (<40)",       len([s for s in scores if s < 40])),
    ("Total Fund Transactions", len(payments)),
    ("Total Documents", len(docs)),
    ("  Mismatch Documents", mm_count),
    ("Total Agencies", len(agencies)),
    ("", ""),
    ("Anomaly Breakdown", ""),
    ("  Cost Overrun",        sum(1 for r in proj_rows if r["Cost Risk"] >= 50)),
    ("  Project Delay",       sum(1 for r in proj_rows if r["Delay Risk"] >= 50)),
    ("  Fin-Phys Mismatch",   sum(1 for r in proj_rows if r["Financial Risk"] >= 50 and r["Physical Progress (%)"] < 25)),
    ("  Duplicate Payment",   sum(1 for r in proj_rows if r["Duplicate Risk"] >= 50)),
    ("  Low Utilization",     sum(1 for r in proj_rows if r["Physical Progress (%)"] < 5)),
    ("  Document Mismatch",   mm_count),
]
df_summary = pd.DataFrame(summary_rows, columns=["Metric", "Count"])
df_proj = pd.DataFrame(proj_rows)
df_ag   = pd.DataFrame(ag_rows)
df_docs = pd.DataFrame(doc_rows)
df_pay  = pd.DataFrame(pay_rows)

HEADER_BG="1A3A5C"; RISK_COLORS={"Critical":"C0392B","High":"E67E22","Medium":"F1C40F","Low":"27AE60"}
thin = Side(style="thin", color="CCCCCC"); border = Border(left=thin,right=thin,top=thin,bottom=thin)

def style_sheet(ws, risk_col=None, mm_col=None):
    for cell in ws[1]:
        cell.font=Font(bold=True,color="FFFFFF",size=11); cell.fill=PatternFill(fill_type="solid",fgColor=HEADER_BG)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=border
    ri=None; mi=None
    for i,c in enumerate(ws[1],1):
        if risk_col and c.value==risk_col: ri=i
        if mm_col   and c.value==mm_col:  mi=i
    for ri2,row in enumerate(ws.iter_rows(min_row=2,max_row=ws.max_row),2):
        for cell in row:
            cell.border=border; cell.alignment=Alignment(vertical="center")
        if ri2%2==0:
            for cell in row:
                if not cell.fill or cell.fill.fgColor.rgb in ("00000000","FFFFFFFF",""):
                    cell.fill=PatternFill(fill_type="solid",fgColor="EAF2FF")
    if ri:
        for row in ws.iter_rows(min_row=2,max_row=ws.max_row,min_col=ri,max_col=ri):
            for cell in row:
                col=RISK_COLORS.get(cell.value)
                if col:
                    cell.fill=PatternFill(fill_type="solid",fgColor=col)
                    cell.font=Font(bold=True,color="FFFFFF"); cell.alignment=Alignment(horizontal="center",vertical="center")
    if mi:
        for row in ws.iter_rows(min_row=2,max_row=ws.max_row,min_col=mi,max_col=mi):
            for cell in row:
                if str(cell.value)=="YES":
                    cell.fill=PatternFill(fill_type="solid",fgColor="922B21")
                    cell.font=Font(bold=True,color="FFFFFF"); cell.alignment=Alignment(horizontal="center")
    for col in ws.columns:
        ml=max((len(str(c.value or "")) for c in col),default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width=min(ml+3,55)
    ws.freeze_panes="A2"

os.makedirs("Dataset",exist_ok=True)
out="Dataset/MPLADS_Sentinel_Showcase_Dataset.xlsx"
with pd.ExcelWriter(out,engine="openpyxl") as writer:
    df_summary.to_excel(writer,sheet_name="Summary",index=False)
    df_proj.to_excel(   writer,sheet_name="Projects",index=False)
    df_ag.to_excel(     writer,sheet_name="Agencies",index=False)
    df_docs.to_excel(   writer,sheet_name="Documents",index=False)
    df_pay.to_excel(    writer,sheet_name="Fund Transactions",index=False)
    wb=writer.book
    style_sheet(wb["Summary"])
    style_sheet(wb["Projects"],risk_col="Risk Level")
    style_sheet(wb["Agencies"],risk_col="Risk Level")
    style_sheet(wb["Documents"],mm_col="Mismatch Flag")
    style_sheet(wb["Fund Transactions"])

print(f"Excel saved: {out}")
print(f"Rows -- Projects:{len(df_proj)}, Agencies:{len(df_ag)}, Docs:{len(df_docs)}, Txns:{len(df_pay)}")
